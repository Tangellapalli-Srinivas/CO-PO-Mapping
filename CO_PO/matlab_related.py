import gc
import logging
import pathlib
import typing
import matlab.engine
import matlab
from threading import Thread, Lock
import openpyxl.utils.exceptions
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils.cell import coordinate_to_tuple
import io
from CO_PO.misc import gen_template


def starting_table_number(table: Table):
    return coordinate_to_tuple(table.ref.split(":")[0])


def parse_tables(tables_to_consider: int, workbook: Workbook, *sheets: typing.List[str]):
    """
    Parse the tables in the workbook.

    :param tables_to_consider: The Top n tables to consider.
    :param workbook: The workbook to parse.
    :param sheets: The sheets to parse.
    :return: The parsed tables.
    """

    for SHEET in sheets:
        _sheet = workbook[SHEET]
        _tables = sorted(_sheet.tables.values(), key=lambda _: starting_table_number(_))

        if len(_tables) < tables_to_consider:
            raise ValueError(
                f"Sheet: {SHEET} has less than {tables_to_consider} tables, Remember"
                f"first {tables_to_consider} tables are considered\nPlease refer Excel Input Format in 'How To' Tab "
            )

        for table in _tables[:tables_to_consider]:
            table: Table = table
            yield [
                [CELL.value for CELL in row]
                for index, row in enumerate(_sheet[table.ref]) if index != 0
            ]


class Engine:
    def __init__(self):
        super().__init__()
        self.matlab_engine: typing.Optional[matlab.engine.MatlabEngine] = None
        self.processing = Lock()
        self.loading = Lock()
        self.raw = []
        self.passed = False

    def is_engine_loading(self):
        return self.loading.locked()

    def is_engine_busy(self):
        return self.processing.locked()

    def start_engine(self) -> bool:
        if self.loading.locked():
            logging.warning("Engine was already loading")
            return False

        with self.loading:
            logging.info("Loading Matlab Engine...")
            self.matlab_engine = matlab.engine.start_matlab()
            self.matlab_engine.cd(str(pathlib.Path(__file__).parent / "Scripts"))
            logging.info("Matlab is now ready")

        return True  # means we have actually loaded engine

    def stop_engine(self):
        logging.warning("Asked to stop the engine")

        self.matlab_engine.quit() if self.matlab_engine else ...
        del self.matlab_engine
        gc.collect()

        self.matlab_engine = None

    def process(self, file_path, exams):
        template = gen_template(False, "Processing another request..., please request after this")
        if self.processing.locked():
            template["passed"] = True
            return False

        with self.processing:
            try:
                self.actual_parse(file_path, int(exams))

                template["status"] = "Successfully Executed your request"
                template["passed"] = True
                self.passed = True

            except openpyxl.utils.exceptions.InvalidFileException:
                template["status"] = "Please save the file as a .xlsx file and try again. Other formats are " \
                                     "not supported "
            except Exception as _:
                template["status"] = "Failed to parse excel file, Reason:\n" + str(_)
                logging.exception("Failed to parse", exc_info=True)

        file_path.unlink()
        return template

    def pure(self):
        if not self.raw:
            return ""
        index = int(self.passed)

        self.raw[index].seek(0)
        return self.raw[index].read()

    def actual_parse(self, saved_excel: pathlib.Path, exams: int):
        workbook = load_workbook(saved_excel, read_only=False)
        logging.info("Workbook loaded")

        if len(workbook.sheetnames) < (exams + 1):
            raise ValueError(
                "There are not enough sheets in the workbook\nWe need X + 1 sheets, where X = Number of Exams\nRefer "
                "Sample Input from the 'How To' Tab "
            )

        sheets = workbook.sheetnames[-(exams + 1):]
        # cot, co_po, weightage, expected = (matlab.double(_) for _ in parse_tables(4, workbook, sheets[0]))
        gc.collect()

        [_.close() for _ in self.raw]
        self.raw.clear()
        self.raw.append(io.StringIO())
        self.raw.append(io.StringIO())

        if not self.matlab_engine:
            self.start_engine()

        self.matlab_engine.bridge(
            *(matlab.double(_) for _ in parse_tables(4, workbook, sheets[0])),
            exams,
            *(matlab.double(_) for _ in parse_tables(2, workbook, *sheets[1:])),
            stdout=self.raw[-1],
            stderr=self.raw[0],
            nargout=0
        )
        workbook.close()

    def close(self):
        [_.close() for _ in self.raw]
        self.matlab_engine.quit()
        del self.matlab_engine
